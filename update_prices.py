#!/usr/bin/env python3
"""
RION Price Updater
Usage: python3 update_prices.py <prices.xlsx>

"価格編集" sheet の価格を読み込み、全ブランチの全HTMLを更新・コミット・プッシュする。
"""

import sys, os, re, subprocess
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.abspath(__file__))

BRANCHES = {
    'UAE': 'main',
    'US':  'us-standalone-site',
    'UK':  'uk-standalone-site',
    'CA':  'ca-standalone-site',
    'NL':  'nl-standalone-site',
    'CH':  'ch-standalone-site',
    'AU':  'au-standalone-site',
}

SERIES_ORDER = ['A','B','C','D','F','G','H','J']
VARIANTS = ['-1','-2','-3','-4']
ALL_CODES = [s+v for s in SERIES_ORDER for v in VARIANTS]

# ── Baseline prices (現在の値・スクリプトが変更を検出するために使う) ──────────
BASELINE = {
    'UAE':{'A-1':17400,'A-2':18000,'A-3':17400,'A-4':17000,'B-1':21400,'B-2':23000,'B-3':21400,'B-4':20000,'C-1':21600,'C-2':23000,'C-3':21600,'C-4':21000,'D-1':26400,'D-2':27600,'D-3':26400,'D-4':26000,'F-1':24400,'F-2':25400,'F-3':24400,'F-4':23000,'G-1':20600,'G-2':21600,'G-3':20600,'G-4':20000,'H-1':21600,'H-2':22400,'H-3':21600,'H-4':21000,'J-1':25600,'J-2':26600,'J-3':25600,'J-4':25000},
    'US': {'A-1':5200,'A-2':5400,'A-3':5200,'A-4':5100,'B-1':6400,'B-2':6900,'B-3':6400,'B-4':6000,'C-1':6500,'C-2':6900,'C-3':6500,'C-4':6200,'D-1':7900,'D-2':8300,'D-3':7900,'D-4':7700,'F-1':7300,'F-2':7600,'F-3':7300,'F-4':6800,'G-1':6200,'G-2':6500,'G-3':6200,'G-4':5900,'H-1':6500,'H-2':6700,'H-3':6500,'H-4':6200,'J-1':7700,'J-2':8000,'J-3':7700,'J-4':7400},
    'UK': {'A-1':4100,'A-2':4300,'A-3':4100,'A-4':4000,'B-1':5100,'B-2':5500,'B-3':5100,'B-4':4700,'C-1':5100,'C-2':5500,'C-3':5100,'C-4':4900,'D-1':6200,'D-2':6600,'D-3':6200,'D-4':6100,'F-1':5800,'F-2':6000,'F-3':5800,'F-4':5400,'G-1':4900,'G-2':5100,'G-3':4900,'G-4':4700,'H-1':5100,'H-2':5300,'H-3':5100,'H-4':4900,'J-1':6100,'J-2':6300,'J-3':6100,'J-4':5800},
    'CA': {'A-1':7100,'A-2':7400,'A-3':7100,'A-4':7000,'B-1':8700,'B-2':9400,'B-3':8700,'B-4':8200,'C-1':8800,'C-2':9400,'C-3':8800,'C-4':8400,'D-1':10700,'D-2':11300,'D-3':10700,'D-4':10500,'F-1':9900,'F-2':10300,'F-3':9900,'F-4':9200,'G-1':8400,'G-2':8800,'G-3':8400,'G-4':8000,'H-1':8800,'H-2':9100,'H-3':8800,'H-4':8400,'J-1':10500,'J-2':10900,'J-3':10500,'J-4':10100},
    'NL': {'A-1':4800,'A-2':5000,'A-3':4800,'A-4':4700,'B-1':5900,'B-2':6300,'B-3':5900,'B-4':5500,'C-1':6000,'C-2':6300,'C-3':6000,'C-4':5700,'D-1':7300,'D-2':7600,'D-3':7300,'D-4':7100,'F-1':6700,'F-2':7000,'F-3':6700,'F-4':6300,'G-1':5700,'G-2':6000,'G-3':5700,'G-4':5400,'H-1':6000,'H-2':6200,'H-3':6000,'H-4':5700,'J-1':7100,'J-2':7400,'J-3':7100,'J-4':6800},
    'CH': {'A-1':4700,'A-2':4900,'A-3':4700,'A-4':4600,'B-1':5800,'B-2':6200,'B-3':5800,'B-4':5400,'C-1':5900,'C-2':6200,'C-3':5900,'C-4':5600,'D-1':7100,'D-2':7500,'D-3':7100,'D-4':6900,'F-1':6600,'F-2':6800,'F-3':6600,'F-4':6100,'G-1':5600,'G-2':5900,'G-3':5600,'G-4':5300,'H-1':5900,'H-2':6000,'H-3':5900,'H-4':5600,'J-1':6900,'J-2':7200,'J-3':6900,'J-4':6700},
    'AU': {'A-1':8100,'A-2':8400,'A-3':8100,'A-4':7900,'B-1':9900,'B-2':10700,'B-3':9900,'B-4':9300,'C-1':9900,'C-2':10500,'C-3':9900,'C-4':9600,'D-1':12100,'D-2':12700,'D-3':12100,'D-4':11900,'F-1':11300,'F-2':11800,'F-3':11300,'F-4':10500,'G-1':9500,'G-2':9900,'G-3':9500,'G-4':9100,'H-1':9900,'H-2':10200,'H-3':9900,'H-4':9600,'J-1':11800,'J-2':12200,'J-3':11900,'J-4':11500},
}

def read_excel_prices(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['価格編集']
    new_prices = {s: {} for s in BRANCHES}
    site_cols = {'UAE':4,'US':5,'UK':6,'CA':7,'NL':8,'CH':9,'AU':10}  # col index (1-based)
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = row[0]
        if not code or code not in ALL_CODES:
            continue
        for site, col in site_cols.items():
            val = row[col - 1]
            if val is not None:
                new_prices[site][code] = int(val)
    return new_prices

# ── Price formatting per site ─────────────────────────────────────────────────

def fmt(site, val):
    """Primary formatted string (used in series pages and most contexts)"""
    v = f'{val:,}'
    return {'UAE': f'AED {v}', 'US': f'${v}', 'UK': f'£{v}',
            'CA': f'CA${v}', 'NL': f'€{v}', 'CH': f'CHF{v}', 'AU': f'${v}'}[site]

def fmt_collection(site, val):
    """Format as it appears in collection.html"""
    v = f'{val:,}'
    if site == 'UAE': return v                # data-price attr (bare number)
    if site == 'US':  return f'${v} USD'
    if site == 'AU':  return f'${v} AUD'
    return fmt(site, val)                     # UK/CA/NL/CH same format

# ── HTML update helpers ───────────────────────────────────────────────────────

def replace_collection_uae(content, code, old_val, new_val):
    """UAE collection.html: replace data-price within data-code context block"""
    old_bare = f'{old_val:,}'
    new_bare = f'{new_val:,}'
    # Match: data-code="A-1" ... data-price="17,400"
    pattern = rf'(data-code="{re.escape(code)}"[^>]*data-price="){re.escape(old_bare)}"'
    return re.sub(pattern, lambda m: m.group(1) + new_bare + '"', content)

def replace_collection_other(content, code, old_fmt, new_fmt):
    """Non-UAE collection.html: replace price in product-price-tag after product-code"""
    # Match: <p class="product-code">B-1</p> ... <p class="product-price-tag">$6,400 USD</p>
    pattern = (rf'(<p class="product-code">{re.escape(code)}</p>)'
               rf'(.*?<p class="product-price-tag">){re.escape(old_fmt)}(</p>)')
    return re.sub(pattern, lambda m: m.group(1) + m.group(2) + new_fmt + m.group(3),
                  content, flags=re.DOTALL)

def replace_all_occurrences(content, old_str, new_str):
    """Simple global replacement (for series pages, index.html etc.)"""
    return content.replace(old_str, new_str)

def update_branch(site, branch, old_prices, new_prices):
    os.chdir(REPO)

    # Stash current work, switch branch
    subprocess.run(['git','stash'], check=False, capture_output=True)
    subprocess.run(['git','checkout', branch], check=True, capture_output=True)

    # Find which prices changed
    changes = {}  # code → (old_val, new_val)
    for code in ALL_CODES:
        old_v = old_prices.get(code)
        new_v = new_prices.get(code)
        if old_v and new_v and old_v != new_v:
            changes[code] = (old_v, new_v)

    if not changes:
        print(f'  {site}: no changes, skipping.')
        subprocess.run(['git','checkout','main'], capture_output=True)
        subprocess.run(['git','stash','pop'], check=False, capture_output=True)
        return

    print(f'  {site}: {len(changes)} price(s) changed')

    # Collect all HTML files
    html_files = [f for f in os.listdir(REPO) if f.endswith('.html')]

    # ── 1. collection.html ─────────────────────────────────────────────
    coll = 'collection.html'
    with open(os.path.join(REPO, coll), 'r', encoding='utf-8') as f:
        content = f.read()

    for code, (old_v, new_v) in changes.items():
        if site == 'UAE':
            content = replace_collection_uae(content, code, old_v, new_v)
        else:
            old_f = fmt_collection(site, old_v)
            new_f = fmt_collection(site, new_v)
            content = replace_collection_other(content, code, old_f, new_f)

    with open(os.path.join(REPO, coll), 'w', encoding='utf-8') as f:
        f.write(content)

    # ── 2. series pages + index.html + other pages ─────────────────────
    # Build replacement pairs for simple string replace
    # Sorted longest-first to avoid substring conflicts
    replacements = []
    for code, (old_v, new_v) in changes.items():
        if site == 'UAE':
            old_str = f'AED {old_v:,}'
            new_str = f'AED {new_v:,}'
            replacements.append((old_str, new_str))
            # Arabic format
            replacements.append((f'من AED {old_v:,}', f'من AED {new_v:,}'))
            # Bare number in data-price (collection already done above)
            # Also handle priceRange: "AED 17,000–"
        elif site == 'US':
            # "$5,200 USD" (collection done) and "$5,200" (series/index)
            replacements.append((f'${old_v:,} USD', f'${new_v:,} USD'))
            replacements.append((f'${old_v:,}', f'${new_v:,}'))
        elif site == 'AU':
            # "$8,100 AUD" (collection done) and "$8,100" (series/index)
            replacements.append((f'${old_v:,} AUD', f'${new_v:,} AUD'))
            replacements.append((f'${old_v:,}', f'${new_v:,}'))
        else:
            # UK/CA/NL/CH: same format everywhere
            replacements.append((fmt(site, old_v), fmt(site, new_v)))

    # Sort: longer strings first to avoid substring replacement issues
    replacements.sort(key=lambda x: -len(x[0]))
    # Deduplicate
    seen = set()
    unique_reps = []
    for pair in replacements:
        if pair[0] not in seen:
            seen.add(pair[0])
            unique_reps.append(pair)

    non_coll = [f for f in html_files if f != 'collection.html']
    for fname in non_coll:
        fpath = os.path.join(REPO, fname)
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        original = content
        for old_str, new_str in unique_reps:
            content = content.replace(old_str, new_str)
        if content != original:
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f'    updated: {fname}')

    # ── 3. Commit & push ──────────────────────────────────────────────
    subprocess.run(['git','add','-u'], check=True, capture_output=True)

    changes_summary = ', '.join(f'{c}:{o}→{n}' for c,(o,n) in list(changes.items())[:5])
    if len(changes) > 5:
        changes_summary += f' (+{len(changes)-5} more)'

    commit_msg = f'Update {site} prices: {changes_summary}'
    result = subprocess.run(
        ['git','commit','-m', commit_msg],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        print(f'    committed: {commit_msg[:80]}')
    else:
        print(f'    nothing to commit for {site}')

    push_result = subprocess.run(
        ['git','push','-u','origin', branch],
        capture_output=True, text=True
    )
    if push_result.returncode == 0:
        print(f'    pushed {branch} ✓')
    else:
        print(f'    push failed: {push_result.stderr[:200]}')

    subprocess.run(['git','checkout','main'], capture_output=True)
    subprocess.run(['git','stash','pop'], check=False, capture_output=True)


def main():
    if len(sys.argv) < 2:
        print('Usage: python3 update_prices.py <prices.xlsx>')
        sys.exit(1)

    xlsx = sys.argv[1]
    if not os.path.exists(xlsx):
        print(f'File not found: {xlsx}')
        sys.exit(1)

    print(f'Reading prices from: {xlsx}')
    new_prices_all = read_excel_prices(xlsx)

    for site, branch in BRANCHES.items():
        print(f'\n[{site}] branch: {branch}')
        update_branch(site, branch, BASELINE[site], new_prices_all[site])

    print('\nDone. Remember to trigger Netlify deploy for each site.')

if __name__ == '__main__':
    main()
